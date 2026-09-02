import os
import io
import shutil
import zipfile
import logging
from pathlib import Path
from datetime import datetime
from typing import List, Tuple, Dict, Any, Optional
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from database.models import SellerRecord

logger = logging.getLogger("amazon_scraper")

COLUMNS = [
    "SUB SUB CATEGORY",
    "SUB SUB SUB CATEGORY",
    "S.NO",
    "Business Name",
    "Business Model",
    "Business Category",
    "Owner Name",
    "Phone Number",
    "Email Address",
    "GST Number",
    "PAN Number",
    "FSSAI Number",
    "Billing Address",
    "x",
    "City",
    "State",
    "Pincode",
    "Country",
    "Website URL",
    "Status",
    "Source"
]

def export_sellers_to_master_excel(
    sellers: List[SellerRecord],
    current_category: str,
    output_path: str = "output/Amazon_Seller_Master_Data.xlsx",
    allow_reprocess: bool = True
) -> Dict[str, Any]:
    """
    Appends or updates category records in the persistent master Excel workbook.
    Strictly preserves all other categories, worksheets, and styling.
    Performs deep backup validation, atomic temporary file replacement,
    post-save verification, and SQLite <-> Excel consistency checking.
    """
    if not current_category or not current_category.strip():
        raise ValueError("Cannot export to Master Excel without a valid category name.")

    norm_category = current_category.strip()
    norm_cat_lower = norm_category.lower()

    final_path = Path(output_path).resolve()
    output_dir = final_path.parent
    output_dir.mkdir(parents=True, exist_ok=True)
    backup_dir = output_dir / "backups"
    backup_dir.mkdir(parents=True, exist_ok=True)

    temp_path = output_dir / f"{final_path.name}.tmp"
    if temp_path.exists():
        try:
            temp_path.unlink()
        except Exception:
            pass

    # Styling definitions
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    # Text format column indices (1-based): Phone(8), GST(10), PAN(11), FSSAI(12), Pincode(17)
    text_cols = {8, 10, 11, 12, 17}

    existing_rows = []
    other_rows = []
    existing_categories = set()
    rows_before = 0
    file_existed = final_path.exists()

    try:
        if file_existed:
            # Step 1: Validate Source Master Before Modification or Backup
            source_size = os.path.getsize(final_path)
            if source_size == 0:
                raise RuntimeError("MASTER EXCEL IS EMPTY OR INVALID (size is 0 bytes).")

            if not zipfile.is_zipfile(str(final_path)):
                raise RuntimeError("MASTER EXCEL IS NOT A VALID XLSX (invalid ZIP structure).")

            with open(final_path, "rb") as f:
                source_bytes = f.read()

            try:
                source_wb = load_workbook(io.BytesIO(source_bytes), data_only=True)
            except Exception as e:
                raise RuntimeError(f"MASTER EXCEL FAILED OPENPYXL VALIDATION: {e}")

            if "Amazon Sellers" not in source_wb.sheetnames:
                source_wb.close()
                raise RuntimeError("MASTER EXCEL IS MISSING 'Amazon Sellers' WORKSHEET.")

            source_ws = source_wb["Amazon Sellers"]
            source_row_count = source_ws.max_row

            # Collect existing rows
            for row in source_ws.iter_rows(min_row=2, values_only=True):
                if row and any(cell is not None for cell in row):
                    cat_val = str(row[0]).strip() if row[0] is not None else ""
                    if cat_val:
                        existing_categories.add(cat_val)
                        existing_rows.append(list(row))
                        if cat_val.lower() != norm_cat_lower:
                            other_rows.append(list(row))
            source_wb.close()
            rows_before = len(existing_rows)

            # Step 2: Lock / Exclusivity Check before mutation
            try:
                with open(final_path, "a+b") as test_f:
                    pass
            except (PermissionError, OSError) as e:
                print("\n========================================")
                print("MASTER EXCEL IS CURRENTLY OPEN/LOCKED.")
                print(f"Please close: {final_path}")
                print("========================================\n")
                raise RuntimeError(f"MASTER EXCEL IS OPEN/LOCKED: {e}")

            # Step 3: Create & Validate Safety Backup
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_path = backup_dir / f"Amazon_Seller_Master_Data_{timestamp}_backup.xlsx"

            try:
                shutil.copy2(str(final_path), str(backup_path))
            except Exception as e:
                raise RuntimeError(f"FAILED TO COPY MASTER TO BACKUP: {e}")

            if not backup_path.exists():
                raise RuntimeError("BACKUP VALIDATION FAILED: Backup file does not exist after copy.")

            backup_size = os.path.getsize(backup_path)
            if backup_size == 0 or not zipfile.is_zipfile(str(backup_path)):
                raise RuntimeError("BACKUP VALIDATION FAILED: Backup file is empty or corrupted ZIP.")

            try:
                with open(backup_path, "rb") as bf:
                    backup_bytes = bf.read()
                backup_wb = load_workbook(io.BytesIO(backup_bytes), data_only=True)
                if "Amazon Sellers" not in backup_wb.sheetnames:
                    backup_wb.close()
                    raise RuntimeError("BACKUP VALIDATION FAILED: Missing 'Amazon Sellers' sheet.")
                backup_row_count = backup_wb["Amazon Sellers"].max_row
                backup_wb.close()
            except Exception as e:
                raise RuntimeError(f"BACKUP VALIDATION FAILED: {e}")

            if backup_row_count != source_row_count:
                raise RuntimeError(f"BACKUP VALIDATION FAILED: Row count mismatch ({source_row_count} vs {backup_row_count})")

            logger.info(f"Created verified safety backup: {backup_path}")

        # Build fresh workbook structure with preserved other categories + new records
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Amazon Sellers"

        # Write Header Row
        ws.append(COLUMNS)
        for col_idx in range(1, len(COLUMNS) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

        ws.row_dimensions[1].height = 28
        ws.freeze_panes = "A2"

        # Append existing preserved records from other categories
        for r in other_rows:
            row_vals = list(r)
            if len(row_vals) < len(COLUMNS):
                row_vals.extend([""] * (len(COLUMNS) - len(row_vals)))
            elif len(row_vals) > len(COLUMNS):
                row_vals = row_vals[:len(COLUMNS)]
            ws.append(row_vals)
            
            row_idx = ws.max_row
            for col_idx in range(1, len(COLUMNS) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                cell.font = Font(name="Calibri", size=10)
                if col_idx in text_cols:
                    cell.number_format = '@'
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif col_idx == 13: # Billing Address
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                elif col_idx == 3: # S.NO
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

        # Step 4: Append current category verified records with clean S.NO 1..N
        added_count = 0
        for idx, s in enumerate(sellers, 1):
            row_data = [
                s.sub_sub_category or norm_category,
                s.sub_sub_sub_category or "",
                idx,  # S.NO 1..N for current category
                s.business_name,
                s.business_model or "Marketplace Seller",
                s.business_category or norm_category,
                s.owner_name or "Not Found",
                str(s.phone_number) if s.phone_number and s.phone_number != "None" else "Not Found",
                s.email_address or "Not Found",
                str(s.gst_number) if s.gst_number and s.gst_number != "None" else "Not Found",
                str(s.pan_number) if s.pan_number and s.pan_number != "None" else "Not Found",
                str(s.fssai_number) if s.fssai_number and s.fssai_number != "None" else "N/A",
                s.billing_address or "Not Found",
                "",  # x column
                s.city or "Not Found",
                s.state or "Not Found",
                str(s.pincode) if s.pincode and s.pincode != "None" else "Not Found",
                s.country or "India",
                s.website_url or "Not Found",
                s.status or "Observed on Amazon",
                s.source or "Amazon"
            ]
            ws.append(row_data)
            added_count += 1

            row_idx = ws.max_row
            for col_idx in range(1, len(COLUMNS) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                cell.font = Font(name="Calibri", size=10)
                if col_idx in text_cols:
                    cell.number_format = '@'
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif col_idx == 13: # Billing Address
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                elif col_idx == 3: # S.NO
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

        # Auto Filter
        ws.auto_filter.ref = ws.dimensions

        # Column widths auto-adjustment
        for col in ws.columns:
            max_len = 0
            col_idx = col[0].column
            col_letter = get_column_letter(col_idx)
            for cell in col:
                val_str = str(cell.value or '')
                if col_idx == 13: # Billing Address
                    max_len = 35
                    break
                max_len = max(max_len, len(val_str))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        # Step 5: Save to Temporary XLSX & explicitly close
        wb.save(str(temp_path))
        wb.close()

        # Step 6: Validate Temporary XLSX before modifying Master
        if not temp_path.exists() or not zipfile.is_zipfile(str(temp_path)):
            raise RuntimeError("Generated temporary Excel file is missing or invalid ZIP.")

        with open(temp_path, "rb") as tf:
            test_wb = load_workbook(tf, data_only=True)
            if "Amazon Sellers" not in test_wb.sheetnames:
                test_wb.close()
                raise RuntimeError("Temporary Excel missing 'Amazon Sellers' worksheet.")
            temp_rows = test_wb["Amazon Sellers"].max_row
            test_wb.close()

        expected_total_rows = len(other_rows) + added_count + 1
        if temp_rows < expected_total_rows:
            raise RuntimeError(f"Temporary Excel row count mismatch: Expected >= {expected_total_rows}, got {temp_rows}")

        # Step 7: Atomic Replacement of Master Excel
        os.replace(str(temp_path), str(final_path))

        # Step 8: Deep Read-Back Verification of Master Excel
        if not final_path.exists() or not zipfile.is_zipfile(str(final_path)):
            raise RuntimeError("Master Excel missing or invalid after atomic replacement.")

        final_size = os.path.getsize(final_path)
        check_wb = load_workbook(str(final_path), data_only=True)
        if "Amazon Sellers" not in check_wb.sheetnames:
            check_wb.close()
            raise RuntimeError("Master Excel missing 'Amazon Sellers' sheet after replacement.")

        check_ws = check_wb["Amazon Sellers"]
        rows_after = check_ws.max_row - 1  # exclude header

        master_categories = set()
        cat_records_in_excel = []
        for row in check_ws.iter_rows(min_row=2, values_only=True):
            if row and any(c is not None for c in row):
                cat_val = str(row[0]).strip() if row[0] is not None else ""
                if cat_val:
                    master_categories.add(cat_val)
                    if cat_val.lower() == norm_cat_lower:
                        cat_records_in_excel.append(row)

        # Step 9: Verify Headers & Cumulative Structure
        header_row = [cell for cell in next(check_ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        if header_row != COLUMNS:
            check_wb.close()
            raise RuntimeError("Master Excel headers do not match expected schema.")

        # Step 10: Verify previous categories still exist
        other_category_names = {str(r[0]).strip() for r in other_rows if r and r[0]}
        if not other_category_names.issubset(master_categories):
            missing_prev = other_category_names - master_categories
            check_wb.close()
            raise RuntimeError(f"EXCEL PERSISTENCE FAILURE: Previous categories lost: {missing_prev}")

        # Step 11: Verify current category exists
        if len(sellers) > 0 and not any(c.lower() == norm_cat_lower for c in master_categories):
            check_wb.close()
            raise RuntimeError(f"EXCEL PERSISTENCE FAILURE: Current completed category '{norm_category}' missing from Master Excel.")

        check_wb.close()

        # Step 12: Verify Current Category Records in Excel
        cat_count_in_excel = len(cat_records_in_excel)
        if len(sellers) > 0 and cat_count_in_excel == 0:
            raise RuntimeError(f"EXCEL PERSISTENCE FAILURE: Expected {len(sellers)} records in Excel for '{norm_category}', found 0.")

        if cat_count_in_excel != len(sellers):
            raise RuntimeError(f"EXCEL PERSISTENCE FAILURE: Category record count mismatch for '{norm_category}' (DB: {len(sellers)}, Excel: {cat_count_in_excel})")

        # Formatted AMAZON EXCEL CHECKPOINT Banner
        print("\n========================================")
        print("AMAZON EXCEL CHECKPOINT")
        print("========================================")
        print(f"\nCategory:\n{norm_category}\n")
        print(f"Excel:\n{output_path}\n")
        print(f"Rows:\n{rows_after}\n")
        print("Validation:\nPASSED\n")
        print("Checkpoint:\nSAVED\n")
        print("========================================\n")

        # Database <-> Excel Consistency Verification Banner
        consistency_passed = (len(sellers) == cat_count_in_excel)
        print(f"DATABASE RECORDS: {len(sellers)}")
        print(f"EXCEL RECORDS: {cat_count_in_excel}")
        print(f"CONSISTENCY: {'PASSED' if consistency_passed else 'FAILED'}\n")

        if not consistency_passed:
            raise RuntimeError(f"DATABASE <-> EXCEL CONSISTENCY FAILED for category '{norm_category}'")

        logger.info(f"Master Excel updated successfully: '{norm_category}' added {added_count} rows. Total rows={rows_after}")

        return {
            "status": "SUCCESS",
            "file_path": str(final_path),
            "rows_before": rows_before,
            "rows_after": rows_after,
            "added_count": added_count,
            "category_records": cat_count_in_excel,
            "total_categories": len(master_categories),
            "file_size": final_size
        }

    except Exception as e:
        logger.exception(f"MASTER EXCEL SAVE FAILED: {e}")
        if temp_path.exists():
            try:
                temp_path.unlink()
            except Exception:
                pass
        print("\n========================================")
        print("MASTER EXCEL SAVE FAILED")
        print(f"Error: {e}")
        print("========================================\n")
        raise
