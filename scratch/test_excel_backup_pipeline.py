import os
import sys
import tempfile
import zipfile
from pathlib import Path
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).parent.parent))

from database.models import SellerRecord
from export.excel_exporter import export_sellers_to_master_excel

def test_excel_backup_pipeline():
    print("==================================================")
    print("STARTING EXCEL BACKUP & LOCK PIPELINE TESTS")
    print("==================================================")

    temp_dir = tempfile.mkdtemp()
    test_excel = os.path.join(temp_dir, "test_master.xlsx")

    # 1. Initial Category Export (Men's Shoes)
    men_sellers = [
        SellerRecord(sub_sub_category="Men's Shoes", s_no=1, business_name="Campusactivewear", phone_number="+919876543210"),
        SellerRecord(sub_sub_category="Men's Shoes", s_no=2, business_name="Cocoblu Retail", phone_number="+915600712026"),
    ]
    res1 = export_sellers_to_master_excel(men_sellers, "Men's Shoes", test_excel, allow_reprocess=False)
    assert res1["status"] == "SUCCESS"
    assert res1["added_count"] == 2
    print("PASS: Initial Master Excel creation verified.")

    # 2. Add Second Category (School Shoes) - Triggers Backup Creation & Verification
    school_sellers = [
        SellerRecord(sub_sub_category="School Shoes", s_no=1, business_name="Liberty Shoes Ltd", phone_number="+911122334455"),
        SellerRecord(sub_sub_category="School Shoes", s_no=2, business_name="Bata India", phone_number="+919988776655"),
    ]
    res2 = export_sellers_to_master_excel(school_sellers, "School Shoes", test_excel, allow_reprocess=False)
    assert res2["status"] == "APPENDED SUCCESSFULLY"
    assert res2["added_count"] == 2

    # Verify backup was created and valid
    backup_dir = os.path.join(temp_dir, "backups")
    backups = list(Path(backup_dir).glob("*.xlsx"))
    assert len(backups) == 1, f"Expected 1 backup file, found {len(backups)}"
    backup_file = backups[0]
    
    # Check backup file validity and row count
    assert backup_file.stat().st_size > 0
    assert zipfile.is_zipfile(backup_file)
    wb_backup = load_workbook(backup_file, data_only=True)
    assert "Amazon Sellers" in wb_backup.sheetnames
    ws_backup = wb_backup["Amazon Sellers"]
    # Source had header + 2 men's shoes rows = 3 rows
    assert ws_backup.max_row == 3, f"Expected 3 rows in backup, got {ws_backup.max_row}"
    wb_backup.close()
    print("PASS: Safety backup correctly copied previous state (3 rows) and passed deep verification.")

    # 3. Add Third Category (Women's Flats Amazon)
    flats_sellers = [
        SellerRecord(sub_sub_category="Women's Flats Amazon", s_no=1, business_name="Yuvi Royals", phone_number="+915853442026"),
        SellerRecord(sub_sub_category="Women's Flats Amazon", s_no=2, business_name="JM LOOKS", phone_number="+919299931199"),
        SellerRecord(sub_sub_category="Women's Flats Amazon", s_no=3, business_name="Cocoblu Retail", phone_number="+915600712026"),
    ]
    res3 = export_sellers_to_master_excel(flats_sellers, "Women's Flats Amazon", test_excel, allow_reprocess=True)
    assert res3["status"] == "APPENDED SUCCESSFULLY"
    assert res3["added_count"] == 3

    # Verify Master contains all 3 categories
    wb_final = load_workbook(test_excel, data_only=True)
    ws_final = wb_final["Amazon Sellers"]
    # 1 header + 2 Men's Shoes + 2 School Shoes + 3 Women's Flats = 8 rows
    assert ws_final.max_row == 8, f"Expected 8 rows in final master, got {ws_final.max_row}"
    categories = set(r[0] for r in ws_final.iter_rows(min_row=2, values_only=True) if r and r[0])
    assert categories == {"Men's Shoes", "School Shoes", "Women's Flats Amazon"}
    wb_final.close()
    print("PASS: Master Excel accurately preserved all categories (8 rows total).")

    # 4. Test Empty / Corrupted Master Protection
    corrupt_excel = os.path.join(temp_dir, "corrupt_master.xlsx")
    with open(corrupt_excel, "w") as f:
        f.write("") # 0 bytes
    
    try:
        export_sellers_to_master_excel(flats_sellers, "Women's Flats Amazon", corrupt_excel)
        assert False, "Failed to reject empty 0-byte master Excel!"
    except RuntimeError as e:
        print(f"PASS: Correctly rejected 0-byte corrupt master: {e}")

    print("\n==================================================")
    print("ALL EXCEL BACKUP & VALIDATION TESTS PASSED!")
    print("==================================================")

if __name__ == "__main__":
    test_excel_backup_pipeline()
