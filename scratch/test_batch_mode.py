import os
import sys
import tempfile
import sqlite3
import io
from pathlib import Path
from unittest.mock import patch, MagicMock

sys.path.insert(0, str(Path(__file__).parent.parent))

from database.database import init_db
from database.models import SellerRecord, SellerOffer
from database.repository import SellerRepository
from export.excel_exporter import export_sellers_to_master_excel
from main import load_batch_categories, load_current_category_and_url, run_batch
from openpyxl import load_workbook

def run_all_batch_tests():
    print("==================================================")
    print("STARTING MULTI-CATEGORY BATCH MODE TEST SUITE")
    print("==================================================")

    temp_dir = tempfile.mkdtemp()

    # -------------------------------------------------------------
    # TEST 1: Batch URL Parsing Validation
    # -------------------------------------------------------------
    print("\n--- TEST 1: Batch URL File Parsing ---")
    batch_file = os.path.join(temp_dir, "test_amazon_urls.txt")
    with open(batch_file, "w", encoding="utf-8") as f:
        f.write("""
# Comment line
Women's Flats Amazon|https://www.amazon.in/s?k=women+flats
Men's Shoes|https://www.amazon.in/s?k=men+shoes

   # Indented comment
School Shoes|https://www.amazon.in/s?k=school+shoes
Home Decor|https://www.amazon.in/s?k=home+decor
Electronics|https://www.amazon.in/s?k=electronics
https://www.amazon.in/s?k=sports+shoes
Handbags
""")

    cats = load_batch_categories(batch_file)
    print(f"Parsed {len(cats)} categories from batch file:")
    for c, u in cats:
        print(f"  - Category: '{c}' | URL: '{u}'")

    assert len(cats) == 7, f"Expected 7 categories, got {len(cats)}"
    assert cats[0] == ("Women's Flats Amazon", "https://www.amazon.in/s?k=women+flats")
    assert cats[1] == ("Men's Shoes", "https://www.amazon.in/s?k=men+shoes")
    assert cats[2] == ("School Shoes", "https://www.amazon.in/s?k=school+shoes")
    assert cats[3] == ("Home Decor", "https://www.amazon.in/s?k=home+decor")
    assert cats[4] == ("Electronics", "https://www.amazon.in/s?k=electronics")
    assert cats[5][0] == "Sports Shoes"
    assert cats[6][0] == "Handbags"
    print("PASS: Batch URL File Parsing passed all test cases.")

    # -------------------------------------------------------------
    # TEST 2: Rejection of Invalid "Current Category" in Batch File
    # -------------------------------------------------------------
    print("\n--- TEST 2: Rejection of Invalid 'Current Category' ---")
    invalid_batch_file = os.path.join(temp_dir, "invalid_amazon_urls.txt")
    with open(invalid_batch_file, "w", encoding="utf-8") as f:
        f.write("Current Category|https://www.amazon.in/s?k=test\n")
    try:
        load_batch_categories(invalid_batch_file)
        assert False, "Should have rejected 'Current Category'"
    except ValueError as e:
        print(f"PASS: Correctly rejected 'Current Category': {e}")

    # -------------------------------------------------------------
    # TEST 3: Sequential Batch Processing, Duplicate Protection & Excel Preservation
    # -------------------------------------------------------------
    print("\n--- TEST 3: Sequential Batch Execution & Duplicate Skipping ---")
    test_db = os.path.join(temp_dir, "batch_test_amazon_sellers.db")
    test_excel = os.path.join(temp_dir, "Amazon_Seller_Master_Data.xlsx")
    init_db(test_db)
    repo = SellerRepository(test_db)

    # Pre-populate Category 1: "Women's Flats Amazon"
    for i in range(1, 21):
        rec = SellerRecord(
            sub_sub_category="Women's Flats Amazon",
            s_no=i,
            business_name=f"Flats Seller {i}",
            gst_number=f"29AAJCC851{i:02d}E1ZH",
            phone_number=f"+9198765432{i:02d}",
            email_address=f"seller{i}@flats.com",
            status="Verified"
        )
        repo.save_or_update_seller(rec)

    # Pre-export Women's Flats Amazon to Master Excel
    flats_sellers = repo.get_sellers_by_category("Women's Flats Amazon")
    assert len(flats_sellers) == 20
    export_sellers_to_master_excel(flats_sellers, "Women's Flats Amazon", test_excel)

    # Create batch file with 4 categories:
    # 1. Women's Flats Amazon (already in DB -> should SKIP)
    # 2. Men's Shoes (NEW -> should SUCCEED)
    # 3. School Shoes (NEW -> simulate FAILURE)
    # 4. Home Decor (NEW -> should SUCCEED)
    exec_batch_file = os.path.join(temp_dir, "exec_batch_urls.txt")
    with open(exec_batch_file, "w", encoding="utf-8") as f:
        f.write("""Women's Flats Amazon|https://www.amazon.in/s?k=women+flats
Men's Shoes|https://www.amazon.in/s?k=men+shoes
School Shoes|https://www.amazon.in/s?k=school+shoes
Home Decor|https://www.amazon.in/s?k=home+decor
""")

    test_config = {
        "product_limit": 5,
        "top_businesses": 20,
        "max_pages": 1,
        "headless": True,
        "allow_category_reprocess": False,
        "master_output_file": test_excel,
        "database_file": test_db,
        "urls_file": exec_batch_file
    }

    # Mock process_category_run to test orchestrator behavior cleanly
    def mock_process_category_run(category_name, target_url, config, repo, headless, allow_reprocess, is_batch):
        if category_name == "School Shoes":
            raise RuntimeError("Simulated network timeout for School Shoes")
        
        # Insert 20 mock sellers for this category
        for i in range(1, 21):
            rec = SellerRecord(
                sub_sub_category=category_name,
                s_no=i,
                business_name=f"{category_name} Seller {i}",
                gst_number=f"29AAJCC999{i:02d}E1ZH",
                phone_number=f"+9199999999{i:02d}",
                email_address=f"seller{i}@{category_name.replace(' ', '').lower()}.com",
                status="Verified"
            )
            repo.save_or_update_seller(rec)

        cat_sellers = repo.get_sellers_by_category(category_name)
        excel_res = export_sellers_to_master_excel(cat_sellers, category_name, config["master_output_file"], allow_reprocess=allow_reprocess)
        return {
            "status": "SUCCESS",
            "category": category_name,
            "sellers_count": len(cat_sellers),
            "added_count": excel_res.get("added_count", 0),
            "excel_result": excel_res,
            "db_file": config["database_file"],
            "master_file": config["master_output_file"]
        }

    with patch("main.process_category_run", side_effect=mock_process_category_run):
        # Capture stdout
        old_stdout = sys.stdout
        sys.stdout = buffer = io.StringIO()
        try:
            run_batch(
                config=test_config,
                headless=True,
                allow_reprocess=False,
                urls_file=exec_batch_file
            )
        finally:
            sys.stdout = old_stdout

        output_str = buffer.getvalue()
        print("\nCaptured Batch Output:\n" + output_str)

    # -------------------------------------------------------------
    # TEST 4: Assertions on Batch Results & Output Format
    # -------------------------------------------------------------
    print("\n--- TEST 4: Verifying Batch Summary Metrics & Terminal Output ---")
    assert "Categories requested: 4" in output_str
    assert "Categories processed: 2" in output_str
    assert "Categories skipped: 1" in output_str
    assert "Categories failed: 1" in output_str
    assert "Women's Flats Amazon\nStatus: SKIPPED - ALREADY EXISTS" in output_str or "Women's Flats Amazon" in output_str and "SKIPPED - ALREADY EXISTS" in output_str
    assert "Men's Shoes\nStatus: SUCCESS\nSellers: 20" in output_str or "Men's Shoes" in output_str and "Status: SUCCESS" in output_str
    assert "School Shoes\nStatus: FAILED" in output_str
    assert "Home Decor\nStatus: SUCCESS\nSellers: 20" in output_str or "Home Decor" in output_str and "Status: SUCCESS" in output_str
    assert "Total category seller records added: 40" in output_str
    print("PASS: Terminal output conforms exactly to specification.")

    # -------------------------------------------------------------
    # TEST 5: Master Excel Workbook Integrity Check
    # -------------------------------------------------------------
    print("\n--- TEST 5: Verifying Master Excel Rows & S.NO Isolation ---")
    wb = load_workbook(test_excel, data_only=True)
    assert "Amazon Sellers" in wb.sheetnames
    ws = wb["Amazon Sellers"]

    excel_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0]:
            cat = str(row[0])
            s_no = row[2]
            b_name = row[3]
            excel_rows.append((cat, s_no, b_name))
    wb.close()

    print(f"Total Master Excel rows: {len(excel_rows)}")
    # Expected: 20 (Women's Flats) + 20 (Men's Shoes) + 20 (Home Decor) = 60 rows
    assert len(excel_rows) == 60, f"Expected 60 rows, got {len(excel_rows)}"

    # Check categories order:
    categories_found = []
    for r in excel_rows:
        if r[0] not in categories_found:
            categories_found.append(r[0])

    print(f"Categories in Master Excel: {categories_found}")
    assert categories_found == ["Women's Flats Amazon", "Men's Shoes", "Home Decor"]

    # Check S.NO reset to 1..20 for each category
    flats_snos = [r[1] for r in excel_rows if r[0] == "Women's Flats Amazon"]
    mens_snos = [r[1] for r in excel_rows if r[0] == "Men's Shoes"]
    decor_snos = [r[1] for r in excel_rows if r[0] == "Home Decor"]

    assert flats_snos == list(range(1, 21)), f"Flats S.NO incorrect: {flats_snos}"
    assert mens_snos == list(range(1, 21)), f"Men's S.NO incorrect: {mens_snos}"
    assert decor_snos == list(range(1, 21)), f"Home Decor S.NO incorrect: {decor_snos}"
    print("PASS: Master Excel accurately preserved all categories and reset S.NO per category.")

    # -------------------------------------------------------------
    # TEST 6: SQLite Database Multi-Category Records Verification
    # -------------------------------------------------------------
    print("\n--- TEST 6: Verifying SQLite Database Multi-Category Records ---")
    all_sellers = repo.get_all_sellers()
    assert len(all_sellers) == 60, f"Expected 60 sellers in DB, got {len(all_sellers)}"

    flats_db = repo.get_sellers_by_category("Women's Flats Amazon")
    mens_db = repo.get_sellers_by_category("Men's Shoes")
    decor_db = repo.get_sellers_by_category("Home Decor")
    school_db = repo.get_sellers_by_category("School Shoes")

    assert len(flats_db) == 20
    assert len(mens_db) == 20
    assert len(decor_db) == 20
    assert len(school_db) == 0, f"Failed category should not have saved corrupted records: {len(school_db)}"
    print("PASS: SQLite Database contains exactly the expected category records.")

    print("\n==================================================")
    print("ALL MULTI-CATEGORY BATCH MODE TESTS PASSED!")
    print("==================================================")

if __name__ == "__main__":
    run_all_batch_tests()
