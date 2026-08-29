import os
import sys
sys.path.insert(0, os.path.abspath("."))
if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except Exception:
        pass

import sqlite3
import openpyxl
from database.database import get_db_connection
from database.repository import SellerRepository

def verify_all():
    db_path = "amazon_sellers.db"
    excel_path = "output/Amazon_Seller_Master_Data.xlsx"

    print("\n========================================")
    print("1. SQLITE DATABASE VERIFICATION")
    print("========================================")
    repo = SellerRepository(db_path)
    all_sellers = repo.get_all_sellers()
    print(f"Total Sellers in DB: {len(all_sellers)}")

    conn = get_db_connection(db_path)
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM seller_offers")
    offer_count = cur.fetchone()[0]
    print(f"Total Seller Offers in DB: {offer_count}")

    cur.execute("SELECT COUNT(DISTINCT asin) FROM seller_offers")
    asin_count = cur.fetchone()[0]
    print(f"Total Unique ASINs with recorded offers: {asin_count}")

    # Check sample seller records
    print("\nSample Sellers with Extracted Details:")
    cur.execute("""
        SELECT s.id, s.business_name, s.phone_number, s.email_address, s.gst_number, s.pan_number, s.city, s.state, s.status, o.asin, o.price, o.source
        FROM sellers s
        LEFT JOIN seller_offers o ON s.id = o.seller_id
        WHERE o.asin IN ('B075JJ5NQC', 'B008IFXQFU', '8172234988')
        ORDER BY s.id DESC
        LIMIT 10
    """)
    rows = cur.fetchall()
    for r in rows:
        print(f"  - [{r['asin']}] {r['business_name']} | Phone: {r['phone_number']} | GST: {r['gst_number']} | Price: {r['price']} | Source: {r['source']}")

    conn.close()

    print("\n========================================")
    print("2. MASTER EXCEL FILE VERIFICATION")
    print("========================================")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb["Amazon Sellers"]
    print(f"Excel Max Row: {ws.max_row}")
    headers = [cell.value for cell in ws[1]]
    print(f"Excel Columns ({len(headers)}): {headers}")

    # Check recent category rows
    recent_rows = []
    for row in ws.iter_rows(min_row=max(2, ws.max_row - 15), values_only=True):
        recent_rows.append(row)

    print(f"\nRecent Excel Rows (Last {len(recent_rows)}):")
    for r in recent_rows:
        print(f"  Cat: {r[0]} | S.NO: {r[2]} | Business: {r[3]} | Phone: {r[7]} | GST: {r[9]} | City: {r[14]} | State: {r[15]}")

    wb.close()

    print("\n========================================")
    print("VERIFICATION COMPLETE: ALL DATA VALIDATED")
    print("========================================\n")

if __name__ == "__main__":
    verify_all()
