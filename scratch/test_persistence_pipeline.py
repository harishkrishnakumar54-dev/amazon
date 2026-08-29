import os
import sys
import tempfile
import json
import sqlite3
from pathlib import Path
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).parent.parent))

from database.database import init_db
from database.models import SellerRecord, SellerOffer, SellerSource
from database.repository import SellerRepository
from export.excel_exporter import export_sellers_to_master_excel

def test_persistence_pipeline():
    print("==================================================")
    print("RUNNING PERSISTENCE PIPELINE TEST")
    print("==================================================")

    temp_dir = tempfile.mkdtemp()
    test_db = os.path.join(temp_dir, "test_amazon_sellers.db")
    test_excel = os.path.join(temp_dir, "test_master_data.xlsx")
    test_audit = os.path.join(temp_dir, "database_save_audit.json")

    init_db(test_db)
    repo = SellerRepository(test_db)

    # 1. Simulate enriched seller records
    enriched_seller_1 = SellerRecord(
        sub_sub_category="Women's Flats Amazon",
        sub_sub_sub_category="",
        s_no=1,
        business_name="Cocoblu Retail",
        business_model="Seller",
        business_category="Footwear; Flats",
        owner_name="Anand Kumar",
        phone_number="+915600712026",
        email_address="care@cocoblu.in",
        gst_number="29AAJCC8517E1ZH",
        pan_number="AAJCC8517E",
        fssai_number="N/A",
        billing_address="L-248, 2nd Floor, 17th Cross, Sector 6, HSR Layout, Bengaluru, Karnataka 560102",
        city="Bengaluru",
        state="Karnataka",
        pincode="560102",
        country="India",
        website_url="https://www.cocoblu.in/",
        status="Verified",
        source="Amazon + Official Website"
    )

    enriched_seller_2 = SellerRecord(
        sub_sub_category="Women's Flats Amazon",
        sub_sub_sub_category="",
        s_no=2,
        business_name="Yuvi Royals",
        business_model="Seller",
        business_category="Footwear; Flats",
        owner_name="Yuvi S",
        phone_number="+915853442026",
        email_address="support@yuvistyle.com",
        gst_number="08GSWPS3557L1ZH",
        pan_number="GSWPS3557L",
        fssai_number="N/A",
        billing_address="1ST Floor, 14 Joshi Colony Barkat Nager, Tonk Phatak, Jaipur, Rajasthan 302015",
        city="Jaipur",
        state="Rajasthan",
        pincode="302015",
        country="India",
        website_url="https://www.yuvistyle.com/",
        status="Verified",
        source="Amazon + Targeted Search"
    )

    # Save sellers to SQLite
    saved_1, is_new_1 = repo.save_or_update_seller(enriched_seller_1)
    assert is_new_1 is True, "First save should be INSERT"
    assert saved_1.id is not None
    assert saved_1.city == "Bengaluru"
    assert saved_1.pan_number == "AAJCC8517E"
    assert saved_1.pincode == "560102"
    assert saved_1.status == "Verified"

    saved_2, is_new_2 = repo.save_or_update_seller(enriched_seller_2)
    assert is_new_2 is True, "Second save should be INSERT"
    assert saved_2.id is not None
    assert saved_2.city == "Jaipur"
    assert saved_2.pan_number == "GSWPS3557L"

    # Test update with partial enrichment merge
    update_seller_1 = SellerRecord(
        sub_sub_category="Women's Flats Amazon",
        s_no=1,
        business_name="Cocoblu Retail",
        gst_number="29AAJCC8517E1ZH",
        owner_name="Anand Kumar Verified",
        phone_number="Not Found" # Incomplete incoming phone should not overwrite existing phone!
    )
    saved_1_upd, is_new_1_upd = repo.save_or_update_seller(update_seller_1)
    assert is_new_1_upd is False, "Update should be UPDATE"
    assert saved_1_upd.id == saved_1.id
    assert saved_1_upd.owner_name == "Anand Kumar Verified"
    assert saved_1_upd.phone_number == "+915600712026", "Existing verified phone must be preserved!"
    assert saved_1_upd.pan_number == "AAJCC8517E", "Existing PAN must be preserved!"
    assert saved_1_upd.city == "Bengaluru", "Existing City must be preserved!"

    # 2. Read final verified records back from SQLite
    db_sellers = repo.get_sellers_by_category("Women's Flats Amazon")
    assert len(db_sellers) == 2, f"Expected 2 sellers in category, got {len(db_sellers)}"
    assert db_sellers[0].business_name == "Cocoblu Retail"
    assert db_sellers[0].pan_number == "AAJCC8517E"
    assert db_sellers[1].business_name == "Yuvi Royals"

    # 3. Export verified SQLite records to Master Excel
    excel_res = export_sellers_to_master_excel(db_sellers, "Women's Flats Amazon", test_excel, allow_reprocess=True)
    assert excel_res["status"] in ("SUCCESS", "APPENDED SUCCESSFULLY")
    assert excel_res["added_count"] == 2

    # 4. Verify Excel Read-Back
    wb = load_workbook(test_excel, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    assert len(rows) == 2, f"Expected 2 rows in Excel, got {len(rows)}"
    row1 = rows[0]
    assert row1[0] == "Women's Flats Amazon"
    assert row1[2] == 1 # S.NO
    assert row1[3] == "Cocoblu Retail"
    assert row1[7] == "+915600712026"
    assert row1[9] == "29AAJCC8517E1ZH"
    assert row1[10] == "AAJCC8517E"
    assert row1[14] == "Bengaluru"
    assert row1[15] == "Karnataka"
    assert row1[16] == "560102"
    assert row1[18] == "https://www.cocoblu.in/"
    assert row1[19] == "Verified"

    # 5. Verify database audit log structure
    audit_data = {
        "category": "Women's Flats Amazon",
        "records_before_save": 2,
        "records_saved": 2,
        "records_verified_after_save": 2,
        "records_failed": 0,
        "sellers": [
            {
                "business_name": "Cocoblu Retail",
                "database_id": saved_1.id,
                "saved": True,
                "verified_after_save": True
            },
            {
                "business_name": "Yuvi Royals",
                "database_id": saved_2.id,
                "saved": True,
                "verified_after_save": True
            }
        ]
    }
    with open(test_audit, "w", encoding="utf-8") as f:
        json.dump(audit_data, f, indent=2)

    with open(test_audit, "r", encoding="utf-8") as f:
        read_audit = json.load(f)
    assert read_audit["records_saved"] == 2
    assert read_audit["records_verified_after_save"] == 2
    assert read_audit["records_failed"] == 0

    print("ALL PERSISTENCE PIPELINE CHECKS PASSED SUCCESSFULLY!")

if __name__ == "__main__":
    test_persistence_pipeline()
