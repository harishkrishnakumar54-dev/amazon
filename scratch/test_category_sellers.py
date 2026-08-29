import os
import sys
import tempfile
import sqlite3
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

from database.database import init_db
from database.models import SellerRecord, SellerOffer, SellerSource
from database.repository import SellerRepository
from export.excel_exporter import export_sellers_to_master_excel
from openpyxl import load_workbook

def run_tests():
    print("==================================================")
    print("STARTING CATEGORY-SPECIFIC SELLER RECORDS TESTS")
    print("==================================================")

    # Use a temporary database for clean, isolated testing
    temp_dir = tempfile.mkdtemp()
    test_db = os.path.join(temp_dir, "test_amazon_sellers.db")
    test_excel = os.path.join(temp_dir, "test_master_data.xlsx")

    init_db(test_db)
    repo = SellerRepository(test_db)

    # -------------------------------------------------------------
    # TEST 1: Cross-Category Separation for Same Business / GST
    # -------------------------------------------------------------
    print("\n--- TEST 1: Inserting Same Business in Multiple Categories ---")
    
    # 1. Insert Cocoblu Retail for Women's Flats
    seller1 = SellerRecord(
        sub_sub_category="Women's Flats",
        s_no=1,
        business_name="Cocoblu Retail",
        gst_number="29AAJCC8517E1ZH",
        phone_number="9876543210",
        email_address="contact@cocoblu.in",
        seller_url="https://amazon.in/shops/cocoblu",
        status="Verified"
    )
    saved1, is_new1 = repo.save_or_update_seller(seller1)
    assert is_new1 is True, "First insertion must be marked as new"
    assert saved1.id is not None, "Saved seller must have an ID"
    print(f"Inserted: ID {saved1.id} | {saved1.sub_sub_category} | {saved1.business_name} | GST: {saved1.gst_number}")

    # Verify count is 1
    sellers = repo.get_all_sellers()
    assert len(sellers) == 1, f"Expected 1 seller, got {len(sellers)}"

    # 2. Insert Cocoblu Retail for Women's Sports Shoes (Same GST & Business Name, Different Category)
    seller2 = SellerRecord(
        sub_sub_category="Women's Sports Shoes",
        s_no=1,
        business_name="Cocoblu Retail",
        gst_number="29AAJCC8517E1ZH",
        phone_number="9876543210",
        email_address="contact@cocoblu.in",
        seller_url="https://amazon.in/shops/cocoblu",
        status="Verified"
    )
    saved2, is_new2 = repo.save_or_update_seller(seller2)
    assert is_new2 is True, "Second category insertion MUST be created as a new category-specific row"
    assert saved2.id != saved1.id, "Second category record MUST have a different ID"
    print(f"Inserted: ID {saved2.id} | {saved2.sub_sub_category} | {saved2.business_name} | GST: {saved2.gst_number}")

    # Verify count is 2
    sellers = repo.get_all_sellers()
    assert len(sellers) == 2, f"Expected 2 sellers, got {len(sellers)}"

    # 3. Insert Cocoblu Retail for Men's Shoes
    seller3 = SellerRecord(
        sub_sub_category="Men's Shoes",
        s_no=1,
        business_name="Cocoblu Retail",
        gst_number="29AAJCC8517E1ZH",
        phone_number="9876543210",
        email_address="contact@cocoblu.in",
        seller_url="https://amazon.in/shops/cocoblu",
        status="Verified"
    )
    saved3, is_new3 = repo.save_or_update_seller(seller3)
    assert is_new3 is True, "Third category insertion MUST be created as a new category-specific row"
    assert saved3.id not in (saved1.id, saved2.id), "Third category record MUST have a unique ID"
    print(f"Inserted: ID {saved3.id} | {saved3.sub_sub_category} | {saved3.business_name} | GST: {saved3.gst_number}")

    # Verify count is 3
    sellers = repo.get_all_sellers()
    assert len(sellers) == 3, f"Expected 3 sellers in DB, got {len(sellers)}"
    print("PASS: Cross-Category Separation (3 distinct category records created).")

    # -------------------------------------------------------------
    # TEST 2: Same Seller + SAME Category -> UPDATE existing record
    # -------------------------------------------------------------
    print("\n--- TEST 2: Same Seller in SAME Category (Update existing row) ---")
    seller2_update = SellerRecord(
        sub_sub_category="Women's Sports Shoes",
        s_no=1,
        business_name="Cocoblu Retail",
        gst_number="29AAJCC8517E1ZH",
        phone_number="9999999999", # Updated phone
        email_address="support@cocoblu.in", # Updated email
        seller_url="https://amazon.in/shops/cocoblu",
        status="Fully Verified"
    )
    saved2_updated, is_new2_upd = repo.save_or_update_seller(seller2_update)
    assert is_new2_upd is False, "Same category re-insertion must NOT create a new row (is_new should be False)"
    assert saved2_updated.id == saved2.id, f"Must update the existing ID {saved2.id}, got {saved2_updated.id}"

    # Verify total count in DB is STILL 3
    sellers_after_upd = repo.get_all_sellers()
    assert len(sellers_after_upd) == 3, f"Expected 3 sellers after update, got {len(sellers_after_upd)}"

    # Verify fields of Women's Sports Shoes updated
    updated_rec = [s for s in sellers_after_upd if s.id == saved2.id][0]
    assert updated_rec.phone_number == "9999999999", "Phone should be updated"
    assert updated_rec.email_address == "support@cocoblu.in", "Email should be updated"

    # Verify Women's Flats was untouched
    flats_rec = [s for s in sellers_after_upd if s.id == saved1.id][0]
    assert flats_rec.phone_number == "9876543210", "Women's Flats phone should be untouched"

    print("PASS: Same Seller + Same Category properly updates existing row without duplicate.")

    # -------------------------------------------------------------
    # TEST 3: Seller Offers Linked to Correct Category Seller Record
    # -------------------------------------------------------------
    print("\n--- TEST 3: Seller Offers Linkage and Deduplication ---")
    
    # Offer for Women's Flats
    repo.add_seller_offer(SellerOffer(
        seller_id=saved1.id,
        asin="B001FLATS",
        product_title="Flat Sandal 1",
        category="Women's Flats",
        seller_name="Cocoblu Retail"
    ))

    # Offer for Women's Sports Shoes
    repo.add_seller_offer(SellerOffer(
        seller_id=saved2.id,
        asin="B002SPORTS",
        product_title="Sports Shoe 1",
        category="Women's Sports Shoes",
        seller_name="Cocoblu Retail"
    ))

    # Add duplicate offer for Women's Sports Shoes
    repo.add_seller_offer(SellerOffer(
        seller_id=saved2.id,
        asin="B002SPORTS",
        product_title="Sports Shoe 1",
        category="Women's Sports Shoes",
        seller_name="Cocoblu Retail"
    ))

    conn = sqlite3.connect(test_db)
    cursor = conn.cursor()
    cursor.execute("SELECT seller_id, asin, category FROM seller_offers ORDER BY id ASC")
    offers = cursor.fetchall()
    conn.close()

    assert len(offers) == 2, f"Expected 2 offers (duplicate ASIN filtered), got {len(offers)}"
    assert offers[0] == (saved1.id, "B001FLATS", "Women's Flats"), f"Offer 1 wrong: {offers[0]}"
    assert offers[1] == (saved2.id, "B002SPORTS", "Women's Sports Shoes"), f"Offer 2 wrong: {offers[1]}"
    print("PASS: Seller offers correctly linked to category seller IDs and deduplicated.")

    # -------------------------------------------------------------
    # TEST 4: Master Excel Export with Same Seller in Multiple Categories
    # -------------------------------------------------------------
    print("\n--- TEST 4: Master Excel Multi-Category Export ---")

    # Export Category 1: Women's Flats
    sellers_flats = repo.get_sellers_by_category("Women's Flats")
    res1 = export_sellers_to_master_excel(sellers_flats, "Women's Flats", test_excel)
    assert res1["status"] == "SUCCESS", f"Export 1 status: {res1['status']}"
    assert res1["added_count"] == 1

    # Export Category 2: Women's Sports Shoes
    sellers_sports = repo.get_sellers_by_category("Women's Sports Shoes")
    res2 = export_sellers_to_master_excel(sellers_sports, "Women's Sports Shoes", test_excel)
    assert res2["status"] == "APPENDED SUCCESSFULLY", f"Export 2 status: {res2['status']}"
    assert res2["added_count"] == 1

    # Export Category 3: Men's Shoes
    sellers_mens = repo.get_sellers_by_category("Men's Shoes")
    res3 = export_sellers_to_master_excel(sellers_mens, "Men's Shoes", test_excel)
    assert res3["status"] == "APPENDED SUCCESSFULLY", f"Export 3 status: {res3['status']}"
    assert res3["added_count"] == 1

    # Read and verify Master Excel contents
    wb = load_workbook(test_excel)
    ws = wb.active

    excel_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0]:
            cat = str(row[0])
            s_no = row[2]
            b_name = row[3]
            excel_rows.append((cat, s_no, b_name))
    wb.close()

    print(f"Excel rows recorded:\n{excel_rows}")
    assert len(excel_rows) == 3, f"Expected 3 rows in Excel, got {len(excel_rows)}"
    
    # Check S.NO resets to 1 per category
    assert excel_rows[0] == ("Women's Flats", 1, "Cocoblu Retail")
    assert excel_rows[1] == ("Women's Sports Shoes", 1, "Cocoblu Retail")
    assert excel_rows[2] == ("Men's Shoes", 1, "Cocoblu Retail")

    print("PASS: Master Excel successfully preserved Cocoblu Retail across all 3 categories with S.NO=1.")

    # -------------------------------------------------------------
    # TEST 5: Direct SQL Verification as specified in requirement 13
    # -------------------------------------------------------------
    print("\n--- TEST 5: Direct SQL Query Verification ---")
    conn = sqlite3.connect(test_db)
    cursor = conn.cursor()
    cursor.execute("""
    SELECT
        id,
        sub_sub_category,
        business_name,
        gst_number
    FROM sellers
    WHERE LOWER(business_name) = LOWER('Cocoblu Retail')
    ORDER BY id ASC;
    """)
    db_results = cursor.fetchall()
    conn.close()

    print("SQL Query Result:")
    for row in db_results:
        print(f"  ID {row[0]}: {row[1]} | {row[2]} | {row[3]}")

    assert len(db_results) == 3, f"Expected 3 rows for Cocoblu Retail, got {len(db_results)}"
    assert db_results[0][1] == "Women's Flats"
    assert db_results[1][1] == "Women's Sports Shoes"
    assert db_results[2][1] == "Men's Shoes"

    print("\n==================================================")
    print("ALL TESTS PASSED SUCCESSFULLY!")
    print("==================================================")

if __name__ == "__main__":
    run_tests()
